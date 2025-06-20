import kivy
from kivy.app import App
from kivy.clock import Clock
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.image import Image
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.dropdown import DropDown
from kivy.uix.popup import Popup
from kivy.uix.scrollview import ScrollView
from kivy.uix.gridlayout import GridLayout
from kivy.graphics import Rectangle, Color
from kivy.graphics.texture import Texture
from kivy.utils import get_color_from_hex

import cv2
import numpy as np
from tensorflow.keras.models import load_model
import csv
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

model = load_model("C:/Users/ALAMAT/Desktop/ass-kivy/food_classifier_model.h5")
class_names = ['hamburger', 'noodles', 'pizza']
prices = {'hamburger': 5.50, 'noodles': 9.50, 'pizza': 8.50}
descriptions = {
    'hamburger': 'Delicious grilled burger with cheese and veggies.',
    'noodles': 'Asian-style noodles with fresh ingredients.',
    'pizza': 'Hot cheesy pizza with tomato sauce and toppings.'
}
images_path = {
    'hamburger': 'C:/Users/ALAMAT/Desktop/ass-kivy/images/hamburger.jpg',
    'noodles': 'C:/Users/ALAMAT/Desktop/ass-kivy/images/noodles.jpg',
    'pizza': 'C:/Users/ALAMAT/Desktop/ass-kivy/images/pizza.jpg'
}
guide_image_path = 'C:/Users/ALAMAT/Desktop/ass-kivy/images/guide.jpg'

class MainLayout(BoxLayout):
    def __init__(self, **kwargs):
        super(MainLayout, self).__init__(**kwargs)
        self.orientation = 'vertical'
        self.spacing = 10
        self.padding = 10
        self.cart = []
        self.qty_value = 1
        self.order_count = self.load_last_order_number() + 1

        with self.canvas.before:
            Color(0.95, 0.97, 1, 1)
            self.rect = Rectangle(size=self.size, pos=self.pos)
        self.bind(size=self._update_rect, pos=self._update_rect)

        self.title_label = Label(text="Just Order", font_size=40, color=get_color_from_hex('#34495e'), size_hint=(1, 0.1), bold=True)
        self.add_widget(self.title_label)

        self.img = Image(source=guide_image_path, size_hint=(1, 0.65))
        self.add_widget(self.img)

        self.result_label = Label(text=" Please select or capture a food", font_size=24, size_hint=(1, 0.05), color=get_color_from_hex('#2c3e50'))
        self.price_label = Label(text=" Price: ", font_size=24, size_hint=(1, 0.05), color=get_color_from_hex('#2c3e50'))
        self.desc_label = Label(text=" Description: ", font_size=24, size_hint=(1, 0.1), color=get_color_from_hex('#2c3e50'))
        self.add_widget(self.result_label)
        self.add_widget(self.price_label)
        self.add_widget(self.desc_label)

        self.dropdown = DropDown()
        self.food_select_btn = Button(text="Choose Food", size_hint=(1, 0.08), background_color=get_color_from_hex('#2980b9'), color=[1, 1, 1, 1])
        self.food_select_btn.bind(on_release=self.dropdown.open)
        self.dropdown.bind(on_select=lambda instance, x: self.set_selected_food(x))

        for food in class_names:
            btn = Button(text=food, size_hint_y=None, height=40)
            btn.bind(on_release=lambda btn: self.dropdown.select(btn.text))
            self.dropdown.add_widget(btn)

        self.selected_food = None
        self.add_widget(self.food_select_btn)

        controls_layout = BoxLayout(size_hint=(1, 0.15), spacing=10)
        qty_label = Label(text="Quantity:", font_size=24, size_hint=(0.3, 1), color=get_color_from_hex('#2c3e50'))
        decrease_btn = Button(text="-", font_size=32, size_hint=(0.2, 1), background_color=get_color_from_hex('#e74c3c'), color=[1, 1, 1, 1])
        decrease_btn.bind(on_press=self.decrease_qty)
        self.qty_display = Label(text=str(self.qty_value), font_size=24, size_hint=(0.2, 1), color=get_color_from_hex('#2c3e50'))
        increase_btn = Button(text="+", font_size=32, size_hint=(0.2, 1), background_color=get_color_from_hex('#2ecc71'), color=[1, 1, 1, 1])
        increase_btn.bind(on_press=self.increase_qty)
        controls_layout.add_widget(qty_label)
        controls_layout.add_widget(decrease_btn)
        controls_layout.add_widget(self.qty_display)
        controls_layout.add_widget(increase_btn)
        self.add_widget(controls_layout)

        self.add_btn = Button(text="Add to Cart", font_size=24, size_hint=(1, 0.1), background_color=get_color_from_hex('#27ae60'), color=[1,1,1,1])
        self.add_btn.bind(on_press=self.add_to_cart)
        self.add_widget(self.add_btn)

        self.action_buttons = BoxLayout(size_hint=(1, 0.1), spacing=10)
        self.confirm_btn = Button(text="Confirm Order", background_color=get_color_from_hex('#16a085'), color=[1,1,1,1])
        self.confirm_btn.bind(on_press=self.show_invoice)
        self.action_buttons.add_widget(self.confirm_btn)
        self.camera_btn = Button(text="Open Camera", background_color=get_color_from_hex('#8e44ad'), color=[1,1,1,1])
        self.camera_btn.bind(on_press=self.start_camera_stream)
        self.action_buttons.add_widget(self.camera_btn)
        self.add_widget(self.action_buttons)

    def _update_rect(self, *args):
        self.rect.pos = self.pos
        self.rect.size = self.size

    def set_selected_food(self, food_name):
        self.selected_food = food_name
        self.food_select_btn.text = f"Selected: {food_name.capitalize()}"
        self.price_label.text = f"Price: {prices[food_name]:.2f} RM"
        self.desc_label.text = f"Description: {descriptions[food_name]}"
        self.result_label.text = f"Food: {food_name}"
        self.img.source = images_path[food_name]

    def increase_qty(self, instance):
        if self.qty_value < 10:
            self.qty_value += 1
            self.qty_display.text = str(self.qty_value)

    def decrease_qty(self, instance):
        if self.qty_value > 1:
            self.qty_value -= 1
            self.qty_display.text = str(self.qty_value)

    def start_camera_stream(self, instance):
        self.capture = cv2.VideoCapture(0)
        self.countdown = 10
        self.countdown_label = Label(text="Capturing in 10s", font_size=24, color=get_color_from_hex('#c0392b'), size_hint=(1, 0.05))
        self.add_widget(self.countdown_label)
        self.frame_event = Clock.schedule_interval(self.update_camera_frame, 1.0 / 30.0)
        self.timer_event = Clock.schedule_interval(self.countdown_tick, 1)

    def update_camera_frame(self, dt):
        ret, frame = self.capture.read()
        if ret:
            buf = cv2.flip(frame, 0).tobytes()
            texture = Texture.create(size=(frame.shape[1], frame.shape[0]), colorfmt='bgr')
            texture.blit_buffer(buf, colorfmt='bgr', bufferfmt='ubyte')
            self.img.texture = texture

    def countdown_tick(self, dt):
        self.countdown -= 1
        if self.countdown_label:
            self.countdown_label.text = f"Capturing in {self.countdown}s"
        if self.countdown <= 0:
            Clock.unschedule(self.frame_event)
            Clock.unschedule(self.timer_event)
            self.capture_image_and_classify()

    def capture_image_and_classify(self):
        ret, frame = self.capture.read()
        self.capture.release()
        self.remove_widget(self.countdown_label)
        if ret:
            img = cv2.resize(frame, (224, 224)) / 255.0
            img = np.expand_dims(img, axis=0)
            prediction = model.predict(img)
            class_idx = np.argmax(prediction)
            confidence = np.max(prediction)
            class_name = class_names[class_idx]
            if confidence >= 0.8:
                self.result_label.text = f"Food: {class_name} ({confidence*100:.2f}%)"
                self.set_selected_food(class_name)
            else:
                self.result_label.text = " No food detected with high confidence."
                self.price_label.text = " Price: "
                self.desc_label.text = " Description: "
                self.food_select_btn.text = "Choose Food"
                self.selected_food = None
                self.img.source = guide_image_path

    def add_to_cart(self, instance):
        if self.selected_food:
            quantity = self.qty_value
            self.cart.append({"food": self.selected_food, "qty": quantity})
            self.result_label.text = f"Added to cart: {quantity} x {self.selected_food}"
            self.selected_food = None
            self.food_select_btn.text = "Choose Food"
            self.price_label.text = " Price: "
            self.desc_label.text = " Description: "
            self.img.source = guide_image_path
            self.qty_value = 1
            self.qty_display.text = str(self.qty_value)
        else:
            self.result_label.text = " Please select a food before adding."

    def show_invoice(self, instance):
        if not self.cart:
            self.result_label.text = "🛒 Your cart is empty."
            return
        if any(item['food'] is None for item in self.cart):
            self.result_label.text = "Invalid item in cart."
            return

        content = BoxLayout(orientation='vertical', spacing=10, padding=10)
        scroll = ScrollView(size_hint=(1, 0.6))
        grid = GridLayout(cols=1, size_hint_y=None)
        grid.bind(minimum_height=grid.setter('height'))

        total = 0
        for item in self.cart:
            name = item['food']
            qty = item['qty']
            subtotal = prices[name] * qty
            total += subtotal
            grid.add_widget(Label(text=f"{qty} x {name.capitalize()} = {subtotal:.2f} RM", font_size=24))

        scroll.add_widget(grid)
        content.add_widget(Label(text=f"Invoice – Order #{self.order_count}", font_size=24, size_hint=(1, 0.1)))
        content.add_widget(scroll)
        content.add_widget(Label(text=f"Total: {total:.2f} RM", font_size=24))

        buttons = BoxLayout(size_hint=(1, 0.2), spacing=10)
        popup = Popup(title="Order Invoice", content=content, size_hint=(0.9, 0.9))

        confirm_btn = Button(text="Confirm & Print", background_color=get_color_from_hex('#27ae60'), color=[1, 1, 1, 1])
        confirm_btn.bind(on_press=lambda x: [self.print_invoice(grid), self.save_order_and_reset(popup)])
        buttons.add_widget(confirm_btn)

        cancel_btn = Button(text="Cancel", background_color=get_color_from_hex('#c0392b'), color=[1, 1, 1, 1])
        cancel_btn.bind(on_press=lambda x: self.reset_app(popup))
        buttons.add_widget(cancel_btn)

        content.add_widget(buttons)
        popup.open()

    def print_invoice(self, grid_widget):
        print("----- Invoice -----")
        for label in reversed(grid_widget.children):
            print(label.text)
        print("-------------------")

    def save_order_and_reset(self, popup):
        filename = "order_history.csv"
        file_exists = os.path.isfile(filename)
        with open(filename, mode='a', newline='') as file:
            writer = csv.writer(file)
            if not file_exists:
                writer.writerow(['Order Number', 'Item', 'Quantity', 'Unit Price', 'Subtotal', 'Date/Time'])
            for item in self.cart:
                name = item['food']
                qty = item['qty']
                unit_price = prices[name]
                subtotal = unit_price * qty
                writer.writerow([self.order_count, name, qty, unit_price, subtotal, datetime.now().strftime('%Y-%m-%d %H:%M:%S')])

        excel_file = "order_history.xlsx"
        if not os.path.exists(excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Orders"
            ws.append(['Order Number', 'Item', 'Quantity', 'Unit Price', 'Subtotal', 'Date/Time'])
        else:
            wb = load_workbook(excel_file)
            ws = wb.active

        for item in self.cart:
            name = item['food']
            qty = item['qty']
            unit_price = prices[name]
            subtotal = unit_price * qty
            ws.append([self.order_count, name, qty, unit_price, subtotal, datetime.now().strftime('%Y-%m-%d %H:%M:%S')])

        wb.save(excel_file)
        self.order_count += 1
        popup.dismiss()
        self.reset_app()
        Popup(title=' Order Confirmed', content=Label(text='Your order has been confirmed thank you.'), size_hint=(0.6, 0.3)).open()

    def load_last_order_number(self):
        filename = "order_history.csv"
        if not os.path.exists(filename):
            return 0
        with open(filename, 'r') as f:
            lines = f.readlines()
            if len(lines) <= 1:
                return 0
            last = lines[-1].split(',')[0]
            return int(last)

    def reset_app(self, popup=None):
        self.cart = []
        self.selected_food = None
        self.result_label.text = "Please select or capture a food"
        self.price_label.text = " Price: "
        self.desc_label.text = " Description: "
        self.food_select_btn.text = "Choose Food"
        self.img.source = guide_image_path
        self.qty_value = 1
        self.qty_display.text = str(self.qty_value)
        if popup:
            popup.dismiss()

class FoodCameraApp(App):
    def build(self):
        return MainLayout()

if __name__ == '__main__':
    FoodCameraApp().run()
